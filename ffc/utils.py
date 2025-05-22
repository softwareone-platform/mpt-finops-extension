import asyncio
import logging
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell

logger = logging.getLogger(__name__)


def find_first(func, iterable, default=None):  # pragma no cover
    return next(filter(func, iterable), default)


def extract_zip_files(file_path, extract_folder) -> tuple[str, list[str]]:
    """
    This function securely zip files from a zip archive into the
    given extract_folder.

    It also prevents Zip Slip attacks by validating that all extracted paths
    remain within the specified extraction directory.

    Args:
        file_path (str): Path to the .zip archive file.
        extract_folder (str): Directory where the archive should be extracted.

    Returns:
        List[str]: List of  filenames located directly in the `extract_folder`
        directory.

    Raises:
        Exception: If a zip entry attempts to escape the extraction directory (Zip Slip).

    """

    if file_path is None or not file_path.lower().endswith(".zip"):
        raise ValueError(f"{file_path} is not a zip archive.")

    with zipfile.ZipFile(file_path, "r") as zip_ref:
        for item in zip_ref.namelist():
            # ensure we are in the right path
            base_path = Path(extract_folder).resolve()
            item_path = (base_path / item).resolve()
            if not item_path.is_relative_to(base_path):
                logger.exception("Zip entry attempts to access another folder.")
                raise Exception(f"WARNING: Unsafe zip entry: {item}")
            zip_ref.extract(item, extract_folder)

    xlsx_file = ""
    json_files = []

    for file in Path(extract_folder).iterdir():
        full_path = file
        if full_path.is_file():
            if full_path.suffix == ".xlsx" and not full_path.name.startswith("~"):
                xlsx_file = str(full_path)
        elif full_path.suffix(".json"):
            json_files.append(str(full_path))
    return xlsx_file, json_files


def read_excel_headers_and_rows(excel_path) -> tuple[list[str], list[tuple[Cell, ...]]]:
    """
    This function reads the headers and data rows from an Excel (.xlsx) file.

    Args:
        excel_path (str): Path to the Excel file to read.

    Returns:
        tuple[list[str], list[tuple]]: A tuple containing a list of header values and a list
        of data rows.

    Raises:
        FileNotFoundError: If the file does not exist or is not accessible.
        InvalidFileException: If the file exists but is not a valid Excel (.xlsx) file.
        ValueError: If the Excel file is empty or missing a header row.
    """

    if not Path(excel_path).is_file():
        raise FileNotFoundError(f"File not found: {excel_path}")
    try:
        workbook = load_workbook(excel_path, read_only=True)
        sheet = workbook.active
        rows = list(
            sheet.iter_rows(min_row=2, values_only=True)
        )  # here there's the actual data
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        return headers, rows
    except zipfile.BadZipFile:
        logger.exception(f"The Excel file is invalid: {excel_path}")
        raise zipfile.BadZipFile(f"The Excel file is invalid: {excel_path}")
    except StopIteration:
        logger.exception("The Excel file is empty or missing header row.")
        raise ValueError("The Excel file is empty or missing header row.")


class AsyncExcelWriter:
    def __init__(self, default_buffer_size: int = 100):
        self.default_buffer_size = default_buffer_size
        self._file_locks: dict[str, asyncio.Lock] = {}
        self._buffers: dict[str, list[Any]] = defaultdict(list)
        self._headers: dict[str, list[str]] = {}

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.flush_all()

    async def add_rows(self, excel_file_path: str, headers: list[str], rows: list[Any]):
        """
        This method adds rows to a Buffer, and it flushes it if
        the contents exceeds the Buffer's size.

        Args:
            excel_file_path (Path): The path to the Excel file to add to the buffer.
            headers (list[str]): The headers to add to the buffer.
            rows (list[Any]): The rows to add to the buffer.
        """
        file_path = str(Path(excel_file_path))

        if file_path not in self._headers:
            self._headers[file_path] = headers
        self._buffers[file_path] = rows

        if len(self._buffers[file_path]) >= self.default_buffer_size:
            # flush contents because they exceed the buffer's size
            await self._flush(Path(file_path))

    async def flush_all(self):
        """
        This method asynchronously flushes all buffered rows
        to their respective Excel files.
        """
        tasks = [
            self._flush(Path(file_path))
            for file_path in self._buffers
            if self._buffers[file_path]
        ]
        await asyncio.gather(*tasks)

    def _get_lock(self, file_path: str):
        """
        This method returns an asyncio lock associated with
        the given file path.

        Args:
            file_path (Path): The path to the file for which a lock is needed.
        """

        if file_path not in self._file_locks:
            self._file_locks[file_path] = asyncio.Lock()
        return self._file_locks[file_path]

    async def _flush(self, file_path: Path):
        """
        This method flushes all buffered rows for a single file to disk.

        Args:
            file_path (Path): The path to the Excel file to be flushed.
        """
        excel_file_path = str(file_path)
        rows_to_write = self._buffers[excel_file_path]
        if not rows_to_write:
            return

        headers = self._headers[excel_file_path]
        lock = self._get_lock(excel_file_path)
        logger.debug(f"Flushing {len(rows_to_write)} rows to {file_path}")

        async with lock:
            await asyncio.get_event_loop().run_in_executor(
                None, self._write_rows, file_path, headers, rows_to_write
            )

        self._buffers[excel_file_path] = []  # empty the buffer

    @staticmethod
    def _write_rows(
        excel_file_path: Path, headers: list[str], rows_to_write: list[Any]
    ):
        """
        This method Adds a row of data to an Excel file.
        It creates the file and writes headers.

        Parameters:
            excel_file_path (str): Path to the Excel file.
            headers (list[str]): List of column headers.
            rows_to_write (list[Any]): A list of rows of data to write.

        Raises:
            IOError: If file cannot be written.
        """

        try:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet()
            ws.append(headers)
            for row in rows_to_write:
                ws.append(row)
            wb.save(str(excel_file_path))
        except Exception as error:
            raise OSError(
                f"Failed to add content to Excel file: {excel_file_path}: {error}"
            )
